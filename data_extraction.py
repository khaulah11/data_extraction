import streamlit as st
import pandas as pd
import numpy as np
import io
import re
import zipfile
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# 1. KONFIGURASI HALAMAN
# ============================================================
st.set_page_config(
    page_title="Modul Ekstraksi Data AAWS | MetMalaysia",
    page_icon="📥",
    layout="wide"
)

# ============================================================
# 2. PENGEPALA APLIKASI
# ============================================================
st.markdown("## 📥 **Sistem Automasi Ekstraksi & Penstrukturan Data AAWS**")
st.caption("Jabatan Meteorologi Malaysia (MetMalaysia) | Pejabat Meteorologi Sabah")
st.divider()

st.markdown("""
Modul ini mengekstrak siri masa hujan daripada fail mentah stesen AAWS, membersihkan rekod tidak sah, 
dan menyusun matriks jadual harian standard mengikut tahun dan stesen.
""")

# ============================================================
# 3. FUNGSI-FUNGSI PEMPROSESAN DIPERTINGKATKAN
# ============================================================
months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

def find_header_row(raw):
    for i in range(len(raw)):
        row = [str(x).strip().lower() if pd.notna(x) else "" for x in raw.iloc[i].tolist()]
        text = " ".join(row)
        if ("year" in text or "tahun" in text) and ("month" in text or "bulan" in text) and ("day" in text or "hari" in text):
            return i
    return None

def clean_filename(name):
    name = str(name).replace(":", "").strip()
    name = re.sub(r'[<>:"/\\|?*]', "_", name)
    name = re.sub(r"\s+", " ", name)
    name = name.strip(" .")
    return name if name != "" else "UNKNOWN_STATION"

def extract_val_from_row(row_series, start_idx):
    """Cari teks pertama yang valid di sebelah kanan (mengatasi isu merged cells)"""
    for k in range(start_idx + 1, len(row_series)):
        v = row_series.iloc[k]
        if pd.notna(v) and str(v).strip() != "" and str(v).strip() != ":":
            val = str(v).strip()
            if val.startswith(":"):
                val = val[1:].strip()
            return val
    return ""

def get_station_info(raw, default_sheet_name=""):
    station, latitude, longitude, elevation = "", "", "", ""
    max_scan_rows = min(15, len(raw)) # Imbas 15 baris teratas sahaja
    
    for i in range(max_scan_rows):
        row_series = raw.iloc[i]
        for j in range(len(row_series)):
            cell_val = row_series.iloc[j]
            if pd.isna(cell_val):
                continue
            text = str(cell_val).strip()
            t_lower = text.lower()
            
            # --- STATION / STESEN ---
            if not station:
                if any(kw in t_lower for kw in ["station name", "nama stesen", "station", "stesen", "stn"]):
                    if ":" in text and not t_lower.endswith(":"):
                        parts = text.split(":", 1)
                        if len(parts) > 1 and parts[1].strip():
                            station = parts[1].strip()
                    if not station:
                        station = extract_val_from_row(row_series, j)
            
            # --- LATITUDE ---
            if not latitude:
                if "lat" in t_lower:
                    if ":" in text and not t_lower.endswith(":"):
                        parts = text.split(":", 1)
                        if len(parts) > 1 and parts[1].strip():
                            latitude = parts[1].strip()
                    if not latitude:
                        latitude = extract_val_from_row(row_series, j)
                        
            # --- LONGITUDE ---
            if not longitude:
                if "long" in t_lower or "lon" in t_lower:
                    if ":" in text and not t_lower.endswith(":"):
                        parts = text.split(":", 1)
                        if len(parts) > 1 and parts[1].strip():
                            longitude = parts[1].strip()
                    if not longitude:
                        longitude = extract_val_from_row(row_series, j)
                        
            # --- ELEVATION ---
            if not elevation:
                if "elev" in t_lower or "alt" in t_lower:
                    if ":" in text and not t_lower.endswith(":"):
                        parts = text.split(":", 1)
                        if len(parts) > 1 and parts[1].strip():
                            elevation = parts[1].strip()
                    if not elevation:
                        elevation = extract_val_from_row(row_series, j)

    # Bersihkan aksara titik bertindih
    station = station.replace(":", "").strip()
    latitude = latitude.replace(":", "").strip()
    longitude = longitude.replace(":", "").strip()
    elevation = elevation.replace(":", "").strip()
    
    # Jika masih gagal kesan, gunakan nama tab sheet
    if not station:
        station = str(default_sheet_name).strip()
        
    return station, latitude, longitude, elevation

def read_station_sheet(excel_file, sheet):
    raw = pd.read_excel(excel_file, sheet_name=sheet, header=None)
    header_row = find_header_row(raw)
    if header_row is None:
        return None, None
    
    info = get_station_info(raw, default_sheet_name=sheet)
    data = raw.iloc[header_row + 1:].copy().iloc[:, :4]
    data.columns = ["Year", "Month", "Day", "Rainfall"]
    
    data["Year"] = pd.to_numeric(data["Year"], errors="coerce")
    data["Month"] = pd.to_numeric(data["Month"], errors="coerce")
    data["Day"] = pd.to_numeric(data["Day"], errors="coerce")
    data["Rainfall"] = pd.to_numeric(data["Rainfall"], errors="coerce")
    
    data = data.dropna(subset=["Year", "Month", "Day"])
    if len(data) == 0:
        return None, info
        
    data["Year"] = data["Year"].astype(int)
    data["Month"] = data["Month"].astype(int)
    data["Day"] = data["Day"].astype(int)
    return data, info

def generate_styled_excel(station_name, station_info, all_data_list):
    data = pd.concat(all_data_list, ignore_index=True)
    data = data.drop_duplicates(subset=["Year", "Month", "Day"], keep="first")
    data = data.sort_values(by=["Year", "Month", "Day"])
    
    output_stream = io.BytesIO()
    
    with pd.ExcelWriter(output_stream, engine="openpyxl") as writer:
        years = sorted(data["Year"].unique())
        for year in years:
            year = int(year)
            data_year = data[data["Year"] == year].copy()
            valid_data = data_year[data_year["Rainfall"].notna() & (data_year["Rainfall"] >= 0)].copy()
            
            table = data_year.pivot(index="Day", columns="Month", values="Rainfall")
            table = table.reindex(columns=range(1, 13)).reindex(range(1, 32))
            table.columns = months
            table.index.name = "hari"
            
            monthly_total = valid_data.groupby("Month")["Rainfall"].sum().reindex(range(1, 13))
            monthly_valid = valid_data.groupby("Month")["Rainfall"].count().reindex(range(1, 13))
            monthly_max = valid_data.groupby("Month")["Rainfall"].max().reindex(range(1, 13))
            
            station, latitude, longitude, elevation = station_info
            clean_station = station.replace(":", "").strip()
            
            info_df = pd.DataFrame({
                0: [
                    "JABATAN METEOROLOGI MALAYSIA",
                    "",
                    "DAILY RAINFALL RECORD IN MILLIMETRES",
                    "",
                    f"STATION  : {clean_station}",
                    f"LATITUDE : {latitude}",
                    f"LONGITUDE: {longitude}",
                    f"ELEVATION: {elevation}",
                    f"YEAR     : {year}"
                ]
            })
            
            info_df.to_excel(writer, sheet_name=str(year), index=False, header=False, startrow=0)
            table.to_excel(writer, sheet_name=str(year), startrow=6)
            
            ws = writer.sheets[str(year)]
            stats_row = 39
            ws.cell(row=stats_row, column=1, value="Total")
            ws.cell(row=stats_row + 1, column=1, value="No. of Valid Data")
            ws.cell(row=stats_row + 2, column=1, value="Max")
            
            for m_num in range(1, 13):
                col_idx = m_num + 1
                tot = monthly_total.loc[m_num]
                ws.cell(row=stats_row, column=col_idx, value=float(tot) if pd.notna(tot) else None)
                
                v_cnt = monthly_valid.loc[m_num]
                ws.cell(row=stats_row + 1, column=col_idx, value=int(v_cnt) if pd.notna(v_cnt) else None)
                
                mx = monthly_max.loc[m_num]
                ws.cell(row=stats_row + 2, column=col_idx, value=float(mx) if pd.notna(mx) else None)
                
    output_stream.seek(0)
    
    # Format openpyxl styles
    wb = load_workbook(output_stream)
    bold_font = Font(bold=True)
    title_font = Font(bold=True, size=14)
    subtitle_font = Font(bold=True, size=12)
    center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    
    for ws in wb.worksheets:
        ws["A1"].font = title_font
        ws["A3"].font = subtitle_font
        
        for cell in ws[7]:
            if cell.column <= 13:
                cell.font = bold_font
                cell.alignment = center
                cell.border = thin_border
                
        for row in ws.iter_rows(min_row=8, max_row=38, min_col=1, max_col=13):
            for cell in row:
                cell.alignment = center
                cell.border = thin_border
                if cell.column >= 2 and isinstance(cell.value, (int, float)):
                    cell.number_format = "0.0"
                    
        stats_row = 39
        for row in ws.iter_rows(min_row=stats_row, max_row=stats_row + 2, min_col=1, max_col=13):
            for cell in row:
                cell.border = thin_border
                cell.alignment = center
                
        ws[f"A{stats_row}"].font = bold_font
        ws[f"A{stats_row + 1}"].font = bold_font
        ws[f"A{stats_row + 2}"].font = bold_font
        
        for col in range(2, 14):
            col_letter = get_column_letter(col)
            ws[f"{col_letter}{stats_row}"].number_format = "0.0"
            ws[f"{col_letter}{stats_row + 1}"].number_format = "0"
            ws[f"{col_letter}{stats_row + 2}"].number_format = "0.0"
            ws.column_dimensions[col_letter].width = 12
            
        ws.column_dimensions["A"].width = 25
        
    final_output = io.BytesIO()
    wb.save(final_output)
    final_output.seek(0)
    return final_output, data

# ============================================================
# 4. RUANG MUAT NAIK & PEMPROSESAN
# ============================================================
uploaded_files = st.file_uploader(
    "📁 Muat naik satu atau lebih fail AAWS (.xlsx / .xls):",
    type=["xlsx", "xls"],
    accept_multiple_files=True
)

if uploaded_files:
    station_groups = {}
    progress_bar = st.progress(0, text="Membaca fail...")
    
    for idx, input_file in enumerate(uploaded_files):
        try:
            excel = pd.ExcelFile(input_file)
            for sheet in excel.sheet_names:
                # Abaikan tab yang bukan stesen cerapan
                if str(sheet).lower().strip() in ['datalist', 'summary', 'senarai', 'sheet1', 'info']:
                    continue
                data_part, info_part = read_station_sheet(excel, sheet)
                if data_part is None:
                    continue
                    
                raw_name = str(info_part[0]).strip() if (info_part and info_part[0]) else str(sheet).strip()
                st_name = raw_name.replace(":", "").strip()
                st_key = re.sub(r"\s+", " ", st_name).strip().upper()
                
                if st_key not in station_groups:
                    station_groups[st_key] = {"name": st_name, "info": info_part, "data": []}
                station_groups[st_key]["data"].append(data_part)
        except Exception as e:
            st.error(f"Ralat pada fail {input_file.name}: {e}")
            
        progress_bar.progress((idx + 1) / len(uploaded_files), text=f"Memproses {input_file.name}...")
        
    progress_bar.empty()
    
    if station_groups:
        st.success(f"✅ Selesai! Sebanyak **{len(station_groups)} stesen** berjaya dikesan & diekstrak.")
        
        # Butang Muat Turun Pukal (.ZIP)
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
            for st_key, st_dict in station_groups.items():
                excel_bytes, _ = generate_styled_excel(st_dict["name"], st_dict["info"], st_dict["data"])
                safe_name = clean_filename(st_dict["name"])
                zip_file.writestr(f"{safe_name}.xlsx", excel_bytes.getvalue())
                
        zip_buffer.seek(0)
        
        st.download_button(
            label="📦 Muat Turun Semua Stesen (.ZIP)",
            data=zip_buffer,
            file_name="Semua_Stesen_Ekstraksi_AAWS.zip",
            mime="application/zip",
            type="primary"
        )
        
        st.write("")
        st.markdown("### 📋 Semakan & Muat Turun Mengikut Stesen")
        
        selected_st_key = st.selectbox(
            "Pilih stesen untuk semakan:",
            options=list(station_groups.keys()),
            format_func=lambda k: station_groups[k]["name"]
        )
        
        curr_st = station_groups[selected_st_key]
        excel_single, combined_df = generate_styled_excel(curr_st["name"], curr_st["info"], curr_st["data"])
        
        clean_st_name = curr_st['name'].replace(":", "").strip()
        
        col_dl, _ = st.columns([2, 5])
        with col_dl:
            st.download_button(
                label="📥 Muat Turun Fail Excel",
                data=excel_single,
                file_name=f"{clean_filename(clean_st_name)}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
        st.dataframe(combined_df.head(25), use_container_width=True)
else:
    st.info("👈 Sila muat naik fail raw AAWS di atas untuk memulakan penstrukturan data.")